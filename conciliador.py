
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rapidfuzz import fuzz
def procesar_conciliacion(
        banco,
        ventas,
        salida,
        omitir_primera_fila=False
):
    # ==========================
    # Columnas
    # ==========================
    
    header_row = 1 if omitir_primera_fila else 0

    columna_banco = "Abono"
    columna_ventas = "Monto"

    columna_banco_concepto = "Concepto / Referencia"
    columna_banco_saldo = "Saldo"
    columna_banco_cfdi = "# CFDI"

    columna_ventas_folio_control = "FOLIO CONTROL"
    columna_ventas_correo = "Correo"
    columna_ventas_nombre = "Nombre"
    columna_ventas_razon_social = "Razon social"

    # Opcional
    columna_ventas_referencia_bancaria = "Referencia Bancaria"

    # ==========================
    # Funciones
    # ==========================

    ## Normaliza un texto eliminando espacios, caracteres especiales y convirtiendo a mayúsculas.

    def normalizar(valor):

        if pd.isna(valor):
            return ""

        texto = str(valor).upper()

        # Eliminar espacios
        texto = re.sub(r"\s+", "", texto)

        # Dejar únicamente letras y números
        texto = re.sub(
            r'[^A-Z0-9]',
            '',
            texto
        )

        return texto



    ## Verifica si un valor está contenido dentro de un concepto, con opción de usar comparación difusa.
    ## - concepto: texto del concepto a analizar
    ## - valor: texto del valor a buscar dentro del concepto
    ## - usar_fuzzy: si es True, se usará comparación difusa para encontrar coincidencias aproximadas
    ## - umbral: porcentaje mínimo de similitud para considerar una coincidencia válida (solo si usar_fuzzy es True)

    def contiene_valor(
        concepto,
        valor,
        usar_fuzzy=False,
        umbral=90
    ):

        concepto_normalizado = normalizar(concepto)
        valor_normalizado = normalizar(valor)

        if not valor_normalizado:
            return False, 0

        # Match exacto
        if valor_normalizado in concepto_normalizado:
            return True, 100

        if not usar_fuzzy:
            return False, 0

        score = fuzz.partial_ratio(
            valor_normalizado,
            concepto_normalizado
        )

        return score >= umbral, score


    # ==========================
    # Lectura
    # ==========================

    print("Leyendo archivos...")

    df_banco = pd.read_excel(
        banco,
        header=header_row
    )

    df_ventas = pd.read_excel(
        ventas,
        header=header_row
    )

    df_banco.columns = df_banco.columns.str.strip()
    df_ventas.columns = df_ventas.columns.str.strip()

    # ==========================
    # Mostrar columnas
    # ==========================

    print("\n=== COLUMNAS BANCO ===")

    for i, columna in enumerate(df_banco.columns, start=1):
        print(f"{i}. [{columna}]")

    print("\n=== COLUMNAS VENTAS ===")

    for i, columna in enumerate(df_ventas.columns, start=1):
        print(f"{i}. [{columna}]")

    # ==========================
    # Validaciones
    # ==========================

    if df_banco.empty:
        raise ValueError(
            f"El archivo {banco} no contiene registros."
        )

    if df_ventas.empty:
        raise ValueError(
            f"El archivo {ventas} no contiene registros."
        )

    columnas_banco_requeridas = [
        columna_banco,
        columna_banco_concepto,
        columna_banco_cfdi,
        columna_banco_saldo
    ]

    columnas_ventas_requeridas = [
        columna_ventas,
        columna_ventas_folio_control,
        columna_ventas_correo,
        columna_ventas_nombre,
        columna_ventas_razon_social
    ]

    faltantes_banco = [
        c
        for c in columnas_banco_requeridas
        if c not in df_banco.columns
    ]

    faltantes_ventas = [
        c
        for c in columnas_ventas_requeridas
        if c not in df_ventas.columns
    ]

    if faltantes_banco:
        raise ValueError(
            f"Columnas faltantes en {banco}: "
            f"{', '.join(faltantes_banco)}"
        )

    if faltantes_ventas:
        raise ValueError(
            f"Columnas faltantes en {ventas}: "
            f"{', '.join(faltantes_ventas)}"
        )

    # ==========================
    # Columna opcional
    # ==========================

    existe_referencia_bancaria = (
        columna_ventas_referencia_bancaria
        in df_ventas.columns
    )

    print(
        f"\nReferencia Bancaria encontrada: "
        f"{existe_referencia_bancaria}"
    )

    # ==========================
    # Montos
    # ==========================

    df_banco[columna_banco] = pd.to_numeric(
        df_banco[columna_banco],
        errors="coerce"
    )

    df_ventas[columna_ventas] = pd.to_numeric(
        df_ventas[columna_ventas],
        errors="coerce"
    )

    # ==========================
    # Nuevas columnas
    # ==========================

    columna_folio_match = "FOLIO_CONTROL_MATCH"
    columna_tipo_match = "TIPO_MATCH"
    columna_valor_match = "VALOR_MATCH"
    columna_score_match = "SCORE_MATCH"

    if columna_folio_match not in df_banco.columns:

        posicion = (
            df_banco.columns.get_loc(
                columna_banco_cfdi
            ) + 1
        )

        df_banco.insert(
            posicion,
            columna_folio_match,
            ""
        )

    if columna_tipo_match not in df_banco.columns:

        posicion = (
            df_banco.columns.get_loc(
                columna_folio_match
            ) + 1
        )

        df_banco.insert(
            posicion,
            columna_tipo_match,
            ""
        )
    if columna_valor_match not in df_banco.columns:

        posicion = (
            df_banco.columns.get_loc(
                columna_tipo_match
            ) + 1
        )

        df_banco.insert(
            posicion,
            columna_valor_match,
            ""
        )

    if columna_score_match not in df_banco.columns:

        posicion = (
            df_banco.columns.get_loc(
                columna_valor_match
            ) + 1
        )

        df_banco.insert(
            posicion,
            columna_score_match,
            pd.NA
        )

    # ==========================
    # Colores
    # ==========================

    COLOR_VERDE = PatternFill(
        fill_type="solid",
        fgColor="92D050"
    )

    COLOR_AMARILLO = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C"
    )

    COLOR_AZUL = PatternFill(
        fill_type="solid",
        fgColor="00B0F0"
    )

    COLOR_ROJO = PatternFill(
        fill_type="solid",
        fgColor="FF0000"
    )

    COLOR_GRIS = PatternFill(
        fill_type="solid",
        fgColor="D9D9D9"
    )

    # ==========================
    # Procesamiento
    # ==========================

    print("\nProcesando registros...")

    colores_filas = {}
    score_match = 0

    for indice_banco, fila_banco in df_banco.iterrows():

        monto_banco = fila_banco[columna_banco]

        coincidencias = df_ventas[
            (df_ventas[columna_ventas] - monto_banco)
            .abs()
            < 0.01
        ]

        color = COLOR_GRIS
        tipo_match = "SIN_MATCH"
        folio_control = ""
        valor_match = ""
        score_match = 0

        if not coincidencias.empty:

            concepto_banco = fila_banco[
                columna_banco_concepto
            ]

            for indice_venta, fila_venta in coincidencias.iterrows():

                # CORREO
                encontro, score = contiene_valor(
                    concepto_banco,
                    fila_venta[columna_ventas_correo],
                    usar_fuzzy=True,
                    umbral=80
                )

                if encontro:

                    color = COLOR_VERDE
                    tipo_match = "CORREO"
                    valor_match = fila_venta[
                        columna_ventas_correo
                    ]
                    score_match = score

                else:

                    # NOMBRE
                    encontro, score = contiene_valor(
                        concepto_banco,
                        fila_venta[columna_ventas_nombre],
                        usar_fuzzy=True,
                        umbral=90
                    )

                    if encontro:

                        color = COLOR_AMARILLO
                        tipo_match = "NOMBRE"
                        valor_match = fila_venta[
                            columna_ventas_nombre
                        ]
                        score_match = score

                    else:

                        # RAZON SOCIAL
                        encontro, score = contiene_valor(
                            concepto_banco,
                            fila_venta[
                                columna_ventas_razon_social
                            ],
                            usar_fuzzy=True,
                            umbral=85
                        )

                        if encontro:

                            color = COLOR_AZUL
                            tipo_match = "RAZON_SOCIAL"
                            valor_match = fila_venta[
                                columna_ventas_razon_social
                            ]
                            score_match = score

                        elif existe_referencia_bancaria:

                            # REFERENCIA BANCARIA
                            encontro, score = contiene_valor(
                                concepto_banco,
                                fila_venta[
                                    columna_ventas_referencia_bancaria
                                ],
                                usar_fuzzy=False
                            )

                            if encontro:

                                color = COLOR_ROJO
                                tipo_match = "REFERENCIA_BANCARIA"
                                valor_match = fila_venta[
                                    columna_ventas_referencia_bancaria
                                ]
                                score_match = score

                            else:
                                continue

                        else:
                            continue

                folio_control = fila_venta[
                    columna_ventas_folio_control
                ]

                break

        df_banco.at[
            indice_banco,
            columna_folio_match
        ] = folio_control

        df_banco.at[
            indice_banco,
            columna_valor_match
        ] = str(valor_match)

        df_banco.at[
            indice_banco,
            columna_tipo_match
        ] = tipo_match

        df_banco.at[
            indice_banco,
            columna_score_match
        ] = float(score_match)

        colores_filas[indice_banco] = color

    # ==========================
    # Guardar
    # ==========================

    df_banco.to_excel(
        salida,
        index=False
    )

    # ==========================
    # Pintar Concepto y columnas nuevas
    # ==========================

    wb = load_workbook(salida)
    ws = wb.active

    # Columna Concepto
    columna_excel_concepto = (
        list(df_banco.columns)
        .index(columna_banco_concepto)
        + 1
    )

    # Columnas nuevas
    columnas_nuevas = [
        columna_folio_match,
        columna_tipo_match,
        columna_valor_match,
        columna_score_match
    ]

    columnas_nuevas_excel = [
        list(df_banco.columns).index(col) + 1
        for col in columnas_nuevas
    ]

    for indice_banco, color in colores_filas.items():

        fila_excel = indice_banco + 2

        # Pintar concepto según el tipo de match
        ws.cell(
            row=fila_excel,
            column=columna_excel_concepto
        ).fill = color

        # Pintar columnas agregadas en amarillo
        for columna_excel in columnas_nuevas_excel:
            ws.cell(
                row=fila_excel,
                column=columna_excel
            ).fill = COLOR_AMARILLO

    # También pintar encabezados de las nuevas columnas
    for columna_excel in columnas_nuevas_excel:
        ws.cell(
            row=1,
            column=columna_excel
        ).fill = COLOR_AMARILLO

    wb.save(salida)

    # ==========================
    # Resumen
    # ==========================

    print("\nProceso terminado.")
    print(f"Archivo generado: {salida}")

    print("\nResumen:")

    print(
        df_banco[columna_tipo_match]
        .value_counts()
    )
    
    return salida