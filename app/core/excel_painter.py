from openpyxl import load_workbook

from app.constants.colors_conciliation import COLOR_AMARILLO
from app.constants.colums_conciliation import (
    COLUMNA_FOLIO_MATCH,
    COLUMNA_TIPO_MATCH,
    COLUMNA_VALOR_MATCH,
    COLUMNA_SCORE_MATCH
)


def pintar_excel(
        salida,
        df_conciliacion,
        colores_filas_banco,
        colores_filas_ventas
):

    wb = load_workbook(salida)

    ws_conciliacion = wb["Conciliacion"]
    ws_ventas = wb["Ventas"]

    # ==========================
    # Hoja Conciliacion
    # ==========================

    columnas_nuevas = [
        COLUMNA_FOLIO_MATCH,
        COLUMNA_TIPO_MATCH,
        COLUMNA_VALOR_MATCH,
        COLUMNA_SCORE_MATCH
    ]

    columnas_nuevas_excel = [
        list(df_conciliacion.columns).index(col) + 1
        for col in columnas_nuevas
    ]

    for indice_banco, color in colores_filas_banco.items():

        fila_excel = indice_banco + 2

        # Pintar fila completa

        for columna_excel in range(
                1,
                ws_conciliacion.max_column + 1
        ):
            ws_conciliacion.cell(
                row=fila_excel,
                column=columna_excel
            ).fill = color

        # Las columnas generadas siguen amarillas

        for columna_excel in columnas_nuevas_excel:

            ws_conciliacion.cell(
                row=fila_excel,
                column=columna_excel
            ).fill = COLOR_AMARILLO

    # Encabezados amarillos

    for columna_excel in columnas_nuevas_excel:

        ws_conciliacion.cell(
            row=1,
            column=columna_excel
        ).fill = COLOR_AMARILLO

    # ==========================
    # Hoja Ventas
    # ==========================

    columna_tipo_match_ventas = (
        ws_ventas.max_column + 1
    )

    # Encabezado

    ws_ventas.cell(
        row=1,
        column=columna_tipo_match_ventas
    ).value = "TIPO_MATCH"

    ws_ventas.cell(
        row=1,
        column=columna_tipo_match_ventas
    ).fill = COLOR_AMARILLO

    # Pintar ventas conciliadas

    for indice_venta, info in colores_filas_ventas.items():

        color = info["color"]
        tipo_match = info["tipo_match"]

        fila_excel = indice_venta + 2

        # Pintar fila completa

        for columna_excel in range(
                1,
                columna_tipo_match_ventas
        ):
            ws_ventas.cell(
                row=fila_excel,
                column=columna_excel
            ).fill = color

        # Escribir tipo match

        ws_ventas.cell(
            row=fila_excel,
            column=columna_tipo_match_ventas
        ).value = tipo_match

        ws_ventas.cell(
            row=fila_excel,
            column=columna_tipo_match_ventas
        ).fill = color

    wb.save(salida)